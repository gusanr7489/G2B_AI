"""제안목차 생성 + 엑셀 변환 서비스"""
import asyncio
import json
import logging
import time as _time
from io import BytesIO

from google import genai
from google.genai import errors as genai_errors
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import get_settings
from prompts.outline_generation import OUTLINE_SYSTEM_PROMPT, build_outline_prompt
from services import progress_store
from services.outline_types import OUTLINE_TYPES, get_label, is_supported

logger = logging.getLogger(__name__)


class UnsupportedProjectTypeError(Exception):
    """ISP/ISMP 등 지원되지 않는 사업 유형."""


def _get_client():
    settings = get_settings()
    return genai.Client(api_key=settings.gemini_api_key)


def _parse_json_response(text: str) -> dict:
    text = text.strip()
    if text.startswith("```"):
        lines = text.split("\n")
        json_lines: list[str] = []
        inside = False
        for line in lines:
            if line.strip().startswith("```") and not inside:
                inside = True
                continue
            if line.strip() == "```" and inside:
                break
            if inside:
                json_lines.append(line)
        text = "\n".join(json_lines)
    return json.loads(text)


# ---------------------------------------------------------------------------
# 생성
# ---------------------------------------------------------------------------


async def generate_outline(bid_id: int, db) -> dict:
    """대상 공고의 제안목차 생성. UnsupportedProjectTypeError 또는 다른 예외 raise."""
    settings = get_settings()

    # 1. 분석 결과 로드
    analysis = await db.fetchrow(
        """
        SELECT id, bid_id, project_type, raw_analysis
        FROM analyses
        WHERE bid_id = $1 AND analysis_status = 'completed'
        ORDER BY created_at DESC LIMIT 1
        """,
        bid_id,
    )
    if not analysis:
        progress_store.emit(bid_id, "완료된 분석 결과가 없습니다", level="error")
        raise ValueError("완료된 분석 결과가 없습니다")

    project_type = analysis["project_type"]
    if not is_supported(project_type):
        labels = ", ".join(cfg["label"] for cfg in OUTLINE_TYPES.values())
        msg = (
            f"{get_label(project_type) or project_type or '미정'} 유형은 "
            f"목차 생성을 지원하지 않습니다 (지원: {labels})"
        )
        progress_store.emit(bid_id, msg, level="error")
        raise UnsupportedProjectTypeError(msg)

    progress_store.emit(bid_id, f"목차 생성 준비 ({get_label(project_type)})")

    raw = analysis["raw_analysis"]
    if isinstance(raw, str):
        analysis_data = json.loads(raw)
    elif isinstance(raw, dict):
        analysis_data = raw
    else:
        analysis_data = {}

    progress_store.emit(bid_id, "샘플 분석본 로드")
    prompt = build_outline_prompt(analysis_data, project_type)

    # 2. Gemini 호출 (모델 fallback 포함)
    client = _get_client()
    gen_config = {
        "system_instruction": OUTLINE_SYSTEM_PROMPT,
        "response_mime_type": "application/json",
        "temperature": 0.2,
    }

    models_to_try = [settings.gemini_model]
    if (
        settings.gemini_fallback_model
        and settings.gemini_fallback_model != settings.gemini_model
    ):
        models_to_try.append(settings.gemini_fallback_model)

    response = None
    last_error: Exception | None = None

    for model_name in models_to_try:
        progress_store.emit(bid_id, f"Gemini 호출 시작 (model={model_name})")
        _t0 = _time.monotonic()
        try:
            for attempt in range(3):
                try:
                    response = await asyncio.to_thread(
                        client.models.generate_content,
                        model=model_name,
                        contents=prompt,
                        config=gen_config,
                    )
                    elapsed = _time.monotonic() - _t0
                    progress_store.emit(
                        bid_id,
                        f"Gemini 응답 수신 ({elapsed:.1f}초, model={model_name})",
                        level="success",
                    )
                    break
                except (genai_errors.ClientError, genai_errors.ServerError) as e:
                    err_str = str(e)
                    if ("429" in err_str or "503" in err_str) and attempt < 2:
                        wait = 20 * (attempt + 1)
                        progress_store.emit(
                            bid_id,
                            f"재시도 대기 {wait}초 (시도 {attempt + 1}/3)",
                            level="warning",
                        )
                        await asyncio.sleep(wait)
                    else:
                        raise
            if response is not None:
                break
        except Exception as e:  # noqa: BLE001
            last_error = e
            if model_name != models_to_try[-1]:
                progress_store.emit(
                    bid_id,
                    f"{model_name} 실패, fallback {models_to_try[-1]}로 전환",
                    level="warning",
                )
                response = None
                continue
            raise

    if response is None:
        raise last_error or RuntimeError("Gemini 응답 없음")

    progress_store.emit(bid_id, "응답 JSON 파싱")
    result = _parse_json_response(response.text)

    # 작성자(assignee)는 비워둠 — 담당자가 엑셀에서 직접 기입.

    # 4. DB UPSERT
    existing = await db.fetchrow(
        "SELECT id FROM proposal_outlines WHERE analysis_id = $1", analysis["id"]
    )
    is_ismp = project_type == "ISMP"
    outline_json = json.dumps(result, ensure_ascii=False)

    if existing:
        await db.execute(
            """
            UPDATE proposal_outlines
            SET outline_data = $1, is_ismp = $2, updated_at = NOW()
            WHERE analysis_id = $3
            """,
            outline_json,
            is_ismp,
            analysis["id"],
        )
    else:
        await db.execute(
            """
            INSERT INTO proposal_outlines (analysis_id, outline_data, is_ismp)
            VALUES ($1, $2, $3)
            """,
            analysis["id"],
            outline_json,
            is_ismp,
        )

    progress_store.emit(bid_id, "DB 저장 완료", level="success")
    logger.info(
        f"제안목차 생성 완료: bid_id={bid_id}, type={project_type}"
    )
    return result


# ---------------------------------------------------------------------------
# 엑셀 빌더
# ---------------------------------------------------------------------------


_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_HEADER_FONT = Font(bold=True)
_TITLE_FONT = Font(bold=True, size=14)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def build_outline_xlsx(outline_data: dict) -> bytes:
    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "Main"
    _build_main_sheet(ws_main, outline_data.get("main") or {})

    ws_outline = wb.create_sheet("제안 목차")
    _build_outline_sheet(ws_outline, outline_data)

    ws_req = wb.create_sheet("RFP 요구사항")
    _build_requirements_sheet(ws_req, outline_data.get("requirements") or [])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_main_sheet(ws, main: dict) -> None:
    rows = [
        ("사업명", main.get("project_name", "")),
        ("고객명", main.get("client", "")),
        ("사업기간", main.get("duration", "")),
        ("용역금액", main.get("amount", "")),
        ("제출일", main.get("submit_date", "")),
        ("제출장소", main.get("submit_place", "")),
        ("비고 / 특이사항", main.get("notes", "")),
    ]
    for i, (label, value) in enumerate(rows, 1):
        c1 = ws.cell(row=i, column=1, value=label)
        c1.font = _HEADER_FONT
        c1.fill = _HEADER_FILL
        c1.alignment = _LEFT
        c1.border = _BORDER
        c2 = ws.cell(row=i, column=2, value=str(value) if value is not None else "")
        c2.alignment = _LEFT
        c2.border = _BORDER
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 80


def _build_outline_sheet(ws, data: dict) -> None:
    main = data.get("main") or {}
    outline = data.get("outline") or []

    ws.cell(row=1, column=1, value=main.get("project_name", "")).font = _TITLE_FONT

    headers = [
        "평가매핑",
        "구분",
        "Level 1",
        "Level 2",
        "Level 3",
        "Level 4",
        "페이지수",
        "작성자",
    ]
    for col, label in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=label)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER
        c.border = _BORDER

    row_num = 5
    for top in outline:
        row_num = _write_section(ws, top, row_num)

    widths = [22, 18, 24, 28, 28, 28, 10, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_section(ws, top: dict, row_num: int) -> int:
    code = top.get("code", "") or ""
    title = top.get("title", "") or ""
    eval_mapping = top.get("eval_mapping") or ""
    page_count = top.get("page_count")

    c_eval = ws.cell(row=row_num, column=1, value=eval_mapping)
    c_eval.alignment = _LEFT
    c_eval.border = _BORDER

    division = f"{code}. {title}".strip(". ")
    c_div = ws.cell(row=row_num, column=2, value=division)
    c_div.font = _HEADER_FONT
    c_div.alignment = _LEFT
    c_div.border = _BORDER

    for col in (3, 4, 5, 6):
        ws.cell(row=row_num, column=col).border = _BORDER

    c_pg = ws.cell(row=row_num, column=7, value=page_count)
    c_pg.alignment = _CENTER
    c_pg.border = _BORDER
    ws.cell(row=row_num, column=8).border = _BORDER

    row_num += 1
    return _write_children(ws, top.get("children") or [], row_num)


def _write_children(ws, nodes: list, row_num: int) -> int:
    for node in nodes:
        level = max(1, min(4, int(node.get("level", 1))))
        number = node.get("number", "") or ""
        title = node.get("title", "") or ""
        text = f"{number} {title}".strip()
        target_col = 2 + level  # L1=3, L2=4, L3=5, L4=6

        c_eval = ws.cell(row=row_num, column=1, value=node.get("eval_mapping") or "")
        c_eval.alignment = _LEFT
        c_eval.border = _BORDER
        ws.cell(row=row_num, column=2).border = _BORDER

        for col in (3, 4, 5, 6):
            cell = ws.cell(row=row_num, column=col)
            cell.border = _BORDER
            if col == target_col:
                cell.value = text
                cell.alignment = _LEFT

        c_pg = ws.cell(row=row_num, column=7, value=node.get("page_count"))
        c_pg.alignment = _CENTER
        c_pg.border = _BORDER
        c_a = ws.cell(row=row_num, column=8, value=node.get("assignee", "") or "")
        c_a.alignment = _CENTER
        c_a.border = _BORDER

        row_num += 1
        children = node.get("children") or []
        if children:
            row_num = _write_children(ws, children, row_num)
    return row_num


def _build_requirements_sheet(ws, requirements: list) -> None:
    headers = ["분류", "CODE", "요구사항 명칭", "정의", "상세설명", "산출정보", "비고"]
    for col, label in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=label)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER
        c.border = _BORDER

    for i, req in enumerate(requirements, 2):
        cells = [
            req.get("category", ""),
            req.get("code", ""),
            req.get("name", ""),
            req.get("definition", ""),
            req.get("detail", ""),
            req.get("output", ""),
            req.get("note", ""),
        ]
        for col, value in enumerate(cells, 1):
            c = ws.cell(row=i, column=col, value=value)
            c.alignment = _LEFT
            c.border = _BORDER

    widths = [18, 14, 28, 30, 50, 22, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
